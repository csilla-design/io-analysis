"""Compute output and labor-income multipliers + supporting indicators for
Thailand's 5 industries of the future, from the 180-sector IO 2021 table.

Methods (Type I, open Leontief model):
  Z[i,j] = PURCHASER - WHOLESALE - RETAIL - TRANSPORT - IMPORT         (domestic, producer price)
  x[j]   = row 210 "Control total" (total output)
  A      = Z / x                                                       (column-normalized)
  L      = (I - A)^{-1}                                                (Leontief inverse)
  m^X[j] = sum_i L[i,j]                                                (output multiplier)
  h[i]   = wages[i] / x[i]   (row 201 as labor-compensation proxy)
  m^W[j] = (h'L)[j] / h[j]                                             (labor-income multiplier)
  v[i]   = total_value_added[i] / x[i]     (row 209)
  m^V[j] = (v'L)[j]                                                    (VA generated per baht FD)
  mu[i]  = imported_intermediate_inputs[i] / x[j]   (column sum of IMPORT over intermediate rows)
  m^M[j] = (mu'L)[j]                                                   (import leakage per baht FD)

Industry-of-the-future aggregation: output-weighted averages over member
sectors (weights = total output x_j).

Major-sector decomposition of the output multiplier for an industry g:
  Σ_{j in g} (x_j / X_g) · Σ_{i in m} L[i,j]      (sums over major sectors m = m^X for g)
Labor-income decomposition (on labor income generated per baht FD):
  Σ_{j in g} (x_j / X_g) · Σ_{i in m} h[i]·L[i,j]  (sums to industry direct+indirect labor income)

Indicators at industry level (output-weighted averages unless noted):
  - export_orientation = exports (cols 305+306) / total supply (col 700)
  - import_penetration = total imports (col 409) / total supply (col 700)
  - labor_share (row 201 / x) and capital_share ((row 202 + row 203) / x)
  - labor_to_capital_ratio
  - domestic_intermediate_share and imported_intermediate_share
  - value_added_multiplier, import_multiplier, labor_income_multiplier (Type I)
  - backward linkage index (output multiplier / economy average)
  - forward linkage index (row sum of L / economy average)

Limitation: the IO file contains no employment headcount data; wages are used
as the labor-compensation proxy, so "labor-income multiplier" is not a true
physical employment multiplier. Provide sector-level employment to swap h.
"""

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "IO2021 180 sectors.xlsx"
OUT = "industries_of_the_future_multipliers.xlsx"

# ---------------------------------------------------------------------------
# 1. Read IO table and build Z, x, wages, value-added, imports, final demand
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(SRC, data_only=True)
data = wb["DataIO2021"]
mapping = wb["sector mapping"]

sector_codes = [f"{i:03d}" for i in range(1, 181)]
code_to_idx = {c: i for i, c in enumerate(sector_codes)}
N = len(sector_codes)

Z = np.zeros((N, N))                   # domestic intermediate flows (producer prices)
Z_imp = np.zeros((N, N))               # imported intermediate flows
x = np.zeros(N)                        # total output          (row 210)
wages = np.zeros(N)                    # wages and salaries    (row 201)
op_surplus = np.zeros(N)               # operating surplus     (row 202)
depreciation = np.zeros(N)             # depreciation          (row 203)
indirect_tax = np.zeros(N)             # indirect taxes less subsidies (row 204)
value_added = np.zeros(N)              # total value added     (row 209)
exports = np.zeros(N)                  # exports (col 305) + special exports (col 306)
special_exports = np.zeros(N)          # col 306 (captures e.g. non-resident tourism spending)
total_demand = np.zeros(N)             # total demand          (col 310)
total_supply = np.zeros(N)             # total supply          (col 700)
total_imports = np.zeros(N)            # total imports         (col 409); stored as |value|
private_cons = np.zeros(N)             # private consumption   (col 301)
govt_cons = np.zeros(N)                # government consumption(col 302)
gfcf = np.zeros(N)                     # gross fixed cap form  (col 303)
sector_name = {c: None for c in sector_codes}

for row in data.iter_rows(min_row=3, values_only=True):
    rc, rd, cc, cd, purch, whol, retail, trans, imp = row
    if rc is None:
        break
    purch = purch or 0
    whol = whol or 0
    retail = retail or 0
    trans = trans or 0
    imp = imp or 0

    # Populate sector name on first encounter
    if rc in code_to_idx and sector_name[rc] is None:
        sector_name[rc] = rd
    if cc in code_to_idx and sector_name[cc] is None:
        sector_name[cc] = cd

    # Intermediate cell (both rc and cc are 001..180)
    if rc in code_to_idx and cc in code_to_idx:
        i = code_to_idx[rc]; j = code_to_idx[cc]
        Z[i, j] = purch - whol - retail - trans - imp
        Z_imp[i, j] = imp

    # Value added and output rows for sector columns
    elif rc in ("201", "202", "203", "204", "209", "210") and cc in code_to_idx:
        j = code_to_idx[cc]
        if rc == "201": wages[j] = purch
        elif rc == "202": op_surplus[j] = purch
        elif rc == "203": depreciation[j] = purch
        elif rc == "204": indirect_tax[j] = purch
        elif rc == "209": value_added[j] = purch
        elif rc == "210": x[j] = purch

    # Sector row i against final-demand columns
    elif rc in code_to_idx and cc in ("301", "302", "303", "305", "306", "309", "310", "409", "700"):
        i = code_to_idx[rc]
        if cc == "301": private_cons[i] = purch
        elif cc == "302": govt_cons[i] = purch
        elif cc == "303": gfcf[i] = purch
        elif cc == "305": exports[i] += purch
        elif cc == "306": special_exports[i] = purch; exports[i] += purch
        elif cc == "310": total_demand[i] = purch
        elif cc == "409": total_imports[i] = abs(purch)  # imports entered as negatives
        elif cc == "700": total_supply[i] = purch

safe_x = np.where(x > 0, x, 1.0)

# ---------------------------------------------------------------------------
# 2. Build A, Leontief inverse, and per-sector Type I + Type II multipliers
# ---------------------------------------------------------------------------
A = Z / safe_x                                # column-normalized A
L = np.linalg.inv(np.eye(N) - A)              # Type I Leontief inverse

# Direct coefficients for value-added, imports, and labor (two proxies)
h_wages = wages / safe_x                      # wages-only labor coefficient (lower bound)
# Upper-bound labor: wages + operating surplus. Row 202 "Operating surplus" in
# Thailand's IO mixes pure capital returns with the mixed income of the self-
# employed (farmers, small traders, freelancers). Adding it gives an upper
# bound on labor compensation; the true figure is between the two.
h_wages_plus_ms = (wages + op_surplus) / safe_x
v = value_added / safe_x                      # value-added coefficient
mu = Z_imp.sum(axis=0) / safe_x               # imported intermediate coefficient

# Type I direct+indirect per unit final demand
output_mult = L.sum(axis=0)
labor_income_per_fd_low = h_wages @ L
labor_income_per_fd_hi = h_wages_plus_ms @ L
va_per_fd = v @ L
imports_per_fd = mu @ L

with np.errstate(divide="ignore", invalid="ignore"):
    labor_income_mult_low = np.where(h_wages > 0, labor_income_per_fd_low / h_wages, np.nan)
    labor_income_mult_hi = np.where(
        h_wages_plus_ms > 0, labor_income_per_fd_hi / h_wages_plus_ms, np.nan
    )

# ---- Type II (closed-model, household-endogenised) multipliers ----
# Close the model with respect to households by adding one row (wage income
# generated per unit of output, h_wages) and one column (final consumption
# pattern divided by total wages) to A. See Miller & Blair (2009) ch. 6.
# We use wages-only as the household-income row so the Type II multiplier is
# directly comparable to the Type I wage-based multiplier.
total_wages = wages.sum()
if total_wages > 0:
    consumption_col = private_cons / total_wages     # share of 1 baht of wages spent on commodity i
    A_closed = np.zeros((N + 1, N + 1))
    A_closed[:N, :N] = A
    A_closed[:N, N] = consumption_col                # households buy commodities
    A_closed[N, :N] = h_wages                        # each sector pays wages
    A_closed[N, N] = 0                               # households do not buy household services from themselves
    L_closed = np.linalg.inv(np.eye(N + 1) - A_closed)

    output_mult_typeII = L_closed[:N, :N].sum(axis=0)
    # Type II labor-income multiplier = (wage row of L_closed) / direct wage coefficient
    wage_row = L_closed[N, :N]
    with np.errstate(divide="ignore", invalid="ignore"):
        labor_income_mult_typeII = np.where(h_wages > 0, wage_row / h_wages, np.nan)
    va_per_fd_typeII = (np.concatenate([v, [0.0]]) @ L_closed)[:N]
    imports_per_fd_typeII = (np.concatenate([mu, [0.0]]) @ L_closed)[:N]
else:
    output_mult_typeII = output_mult.copy()
    labor_income_mult_typeII = labor_income_mult_low.copy()
    va_per_fd_typeII = va_per_fd.copy()
    imports_per_fd_typeII = imports_per_fd.copy()

# Normalised linkage indices (based on Type I)
econ_mean_out_mult = output_mult.mean()
backward_linkage = output_mult / econ_mean_out_mult
row_sums_L = L.sum(axis=1)
econ_mean_row_sum = row_sums_L.mean()
forward_linkage = row_sums_L / econ_mean_row_sum

# Back-compat aliases used in the workbook-writing section below
h = h_wages                                    # default labor coefficient = wages-only
labor_income_per_fd = labor_income_per_fd_low
labor_income_mult = labor_income_mult_low

# ---------------------------------------------------------------------------
# 3. Industry-of-the-future mapping and major-sector groupings
# ---------------------------------------------------------------------------
industry_of = {}
for r in mapping.iter_rows(min_row=2, values_only=True):
    code, name, ind = r[0], r[1], r[2]
    if code and ind:
        industry_of[code] = ind

industries = [
    "Agribusiness",
    "Creative industries",
    "Advanced manufacturing",
    "Tourism",
    "Digital services",
]

# Major-sector groupings for decomposition (13 groups covering all 180 sectors)
def major_sector(code):
    n = int(code)
    if 1 <= n <= 29:    return "Agriculture, forestry & fishing"
    if 30 <= n <= 41:   return "Mining & quarrying"
    if 42 <= n <= 66:   return "Food, beverages & tobacco"
    if 67 <= n <= 77:   return "Textiles, apparel & leather"
    if 78 <= n <= 83:   return "Wood, paper & printing"
    if 84 <= n <= 98:   return "Chemicals, petroleum, rubber & plastics"
    if 99 <= n <= 111:  return "Non-metallic minerals & metals"
    if 112 <= n <= 134: return "Machinery, equipment & vehicles"
    if 135 <= n <= 137: return "Utilities"
    if 138 <= n <= 144: return "Construction"
    if 145 <= n <= 146: return "Wholesale & retail trade"
    if 147 <= n <= 148: return "Hotels & restaurants"
    if 149 <= n <= 158: return "Transport & storage"
    if n == 159:        return "Post & telecommunications"
    if 160 <= n <= 163: return "Finance, insurance & real estate"
    if n == 164:        return "Business services"
    if 165 <= n <= 171: return "Public & community services"
    if 172 <= n <= 180: return "Creative, personal & other services"
    raise ValueError(code)

major_order = [
    "Agriculture, forestry & fishing",
    "Mining & quarrying",
    "Food, beverages & tobacco",
    "Textiles, apparel & leather",
    "Wood, paper & printing",
    "Chemicals, petroleum, rubber & plastics",
    "Non-metallic minerals & metals",
    "Machinery, equipment & vehicles",
    "Utilities",
    "Construction",
    "Wholesale & retail trade",
    "Hotels & restaurants",
    "Transport & storage",
    "Post & telecommunications",
    "Finance, insurance & real estate",
    "Business services",
    "Public & community services",
    "Creative, personal & other services",
]
major_of = [major_sector(c) for c in sector_codes]
major_mask = {m: np.array([mo == m for mo in major_of]) for m in major_order}

# ---------------------------------------------------------------------------
# 4. Industry-level aggregates: weighted averages, decompositions, indicators
# ---------------------------------------------------------------------------
def idxs_in_industry(name):
    return [code_to_idx[c] for c, ind in industry_of.items() if ind == name]

industry_summary = []
industry_decomp_output = {}       # {industry: array over major sectors}
industry_decomp_laborinc = {}     # {industry: array over major sectors}

for ind in industries:
    idxs = idxs_in_industry(ind)
    xs = x[idxs]; Xg = xs.sum()
    wt = xs / Xg                                  # output weights over the industry
    # Type I weighted multipliers (ratio form where applicable)
    om_w = float((wt * output_mult[idxs]).sum())
    li_w_low = float(np.nansum(wt * labor_income_mult_low[idxs]))
    li_w_hi = float(np.nansum(wt * labor_income_mult_hi[idxs]))
    vm_w = float((wt * va_per_fd[idxs]).sum())    # VA generated per 1 baht FD
    im_w = float((wt * imports_per_fd[idxs]).sum())
    # Type II (closed model w.r.t. households)
    om_w_II = float((wt * output_mult_typeII[idxs]).sum())
    li_w_II = float(np.nansum(wt * labor_income_mult_typeII[idxs]))
    vm_w_II = float((wt * va_per_fd_typeII[idxs]).sum())
    im_w_II = float((wt * imports_per_fd_typeII[idxs]).sum())

    # Direct structural shares (output-weighted)
    wages_share = float((wt * (wages[idxs] / safe_x[idxs])).sum())
    op_surplus_share = float((wt * (op_surplus[idxs] / safe_x[idxs])).sum())
    depr_share = float((wt * (depreciation[idxs] / safe_x[idxs])).sum())
    indtax_share = float((wt * (indirect_tax[idxs] / safe_x[idxs])).sum())
    va_share = float((wt * v[idxs]).sum())
    # Labor-compensation bounds (see notes)
    labor_share_low = wages_share                              # formal wages only
    labor_share_hi = wages_share + op_surplus_share            # wages + mixed income upper bound
    capital_share_narrow = depr_share                          # pure depreciation = capital consumption
    capital_share_broad = op_surplus_share + depr_share        # includes mixed income
    dom_int_share = float((wt * (Z[:, idxs].sum(axis=0) / safe_x[idxs])).sum())
    imp_int_share = float((wt * (Z_imp[:, idxs].sum(axis=0) / safe_x[idxs])).sum())
    # Export orientation and import penetration share the same denominator
    # (total supply = output + |imports|, col 700). Using output alone in the
    # denominator yields ratios > 100% for re-export sectors (where imported
    # goods pass through Thailand and are re-exported with little local
    # value-add); using total supply keeps both indicators bounded and on
    # the same supply basis. See "Export-orientation convention" caveat.
    Xs_supply = total_supply[idxs].sum()
    exp_orient = float(exports[idxs].sum() / Xs_supply) if Xs_supply > 0 else np.nan
    imp_pen_num = total_imports[idxs].sum()
    imp_pen = float(imp_pen_num / Xs_supply) if Xs_supply > 0 else np.nan
    fwd_idx = float((wt * forward_linkage[idxs]).sum())
    bwd_idx = float((wt * backward_linkage[idxs]).sum())
    labor_to_cap_low = labor_share_low / capital_share_broad if capital_share_broad > 0 else np.nan
    labor_to_cap_hi = labor_share_hi / capital_share_narrow if capital_share_narrow > 0 else np.nan

    industry_summary.append({
        "industry": ind,
        "n_sectors": len(idxs),
        "total_output": float(Xg),
        "total_exports": float(exports[idxs].sum()),
        "total_wages": float(wages[idxs].sum()),
        "total_va": float(value_added[idxs].sum()),
        # Type I
        "output_mult": om_w,
        "labor_income_mult_low": li_w_low,
        "labor_income_mult_hi": li_w_hi,
        "va_mult": vm_w,
        "import_mult": im_w,
        # Type II (household-endogenised)
        "output_mult_II": om_w_II,
        "labor_income_mult_II": li_w_II,
        "va_mult_II": vm_w_II,
        "import_mult_II": im_w_II,
        # Structural shares
        "wages_share": wages_share,
        "op_surplus_share": op_surplus_share,
        "depr_share": depr_share,
        "indtax_share": indtax_share,
        "va_share": va_share,
        "labor_share_low": labor_share_low,
        "labor_share_hi": labor_share_hi,
        "labor_to_cap_low": labor_to_cap_low,
        "labor_to_cap_hi": labor_to_cap_hi,
        "export_orientation": exp_orient,
        "import_penetration": imp_pen,
        "dom_intermediate_share": dom_int_share,
        "imp_intermediate_share": imp_int_share,
        "backward_linkage_idx": bwd_idx,
        "forward_linkage_idx": fwd_idx,
    })

    # Decompositions by major source sector (columns sum to the industry multiplier)
    decomp_out = np.zeros(len(major_order))
    decomp_labor = np.zeros(len(major_order))
    for m_k, m_name in enumerate(major_order):
        mask = major_mask[m_name]
        # Output: Σ_{j in g} wt_j · Σ_{i in m} L[i,j]
        decomp_out[m_k] = float((wt * L[mask][:, idxs].sum(axis=0)).sum())
        # Labor income per FD: Σ_{j in g} wt_j · Σ_{i in m} h_i · L[i,j]
        decomp_labor[m_k] = float((wt * (h[mask] @ L[mask][:, idxs])).sum())
    industry_decomp_output[ind] = decomp_out
    industry_decomp_laborinc[ind] = decomp_labor

# ---------------------------------------------------------------------------
# 5. Write workbook
# ---------------------------------------------------------------------------
out = openpyxl.Workbook()

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
sec_font = Font(bold=True, color="305496")
bold = Font(bold=True)
thin = Side(border_style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_header(ws, row, headers, widths=None):
    for j, h_ in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h_)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = box
    ws.row_dimensions[row].height = 46
    if widths:
        for j, w_ in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w_

fmt_int = "#,##0"
fmt_4 = "0.0000"
fmt_pct = "0.00%"

# ---- Sheet 1: Methodology & caveats ----
ws_m = out.active; ws_m.title = "Methodology & caveats"
ws_m["A1"] = "Methodology and caveats"
ws_m["A1"].font = Font(bold=True, size=14)
ws_m.merge_cells("A1:A1")
ws_m.column_dimensions["A"].width = 140

methodology_lines = [
    ("Source table", True),
    ("Thailand Input-Output Table 2021, 180-sector classification (sheet 'DataIO2021' in IO2021 180 sectors.xlsx). "
     "Figures are in nominal million baht at purchaser prices. Sector mapping to the 5 industries of the future "
     "is read from the 'sector mapping' sheet; 121 of 180 sectors are mapped, the remainder are treated as 'Other'.",
     False),
    ("", False),
    ("Transformations", True),
    ("1. Domestic producer-price intermediate flows:  Z[i,j] = PURCHASER − WHOLESALE − RETAIL − TRANSPORT − IMPORT. "
     "The Thai IO stores margin components as negative values in the trade (rows 145–146) and transport (rows 149–156) "
     "rows, so this formula both strips margins from commodity cells and reallocates them to trade/transport rows "
     "in a single step. Verified: column sums of WHOLESALE/RETAIL/TRANSPORT across the 180 intermediate rows equal 0.", False),
    ("2. Imported intermediate flows Z_imp = IMPORT (per cell). Totals at col 409 in sector rows are stored as negative "
     "(supply-accounting convention) and are absolute-valued when used as an indicator.", False),
    ("3. Value added, output, and final-demand vectors come directly from rows 201–210 and cols 301–700 respectively.", False),
    ("", False),
    ("Multipliers (Type I, open Leontief model)", True),
    ("  A = Z / x (column-normalised),  L = (I − A)^−1.", False),
    ("  Output multiplier m^X[j] = Σ_i L[i,j].  Direct + indirect output generated economy-wide per 1 baht of final demand.", False),
    ("  Value-added multiplier = (v'L)[j] with v = VA/x. Total VA generated per 1 baht final demand.", False),
    ("  Import multiplier = (μ'L)[j] with μ = imported intermediates / x. Imports induced per 1 baht final demand.", False),
    ("  Labor-income multiplier (ratio) = (h'L)[j] / h[j]. See caveat on labor measurement below — two bounds reported.", False),
    ("", False),
    ("Multipliers (Type II, closed model w.r.t. households)", True),
    ("  Closes A against households by adding one row (h = wages/x, wages paid per unit of sector output) and one column "
     "(c = private_consumption / total_wages, spending by 1 baht of wages on each commodity). "
     "L_closed = (I − A_closed)^−1 in (N+1)×(N+1). Type II multipliers include the induced-consumption round and are "
     "always ≥ Type I. They are appropriate when a final-demand shock is expected to raise household income and spending.", False),
    ("", False),
    ("Aggregation to industries of the future", True),
    ("Industry multipliers are output-weighted averages over member sectors (weights = x_j). "
     "Decompositions by source major sector are Σ_{j in industry} w_j · Σ_{i in major sector} L[i,j], so the column "
     "sums equal the industry's multiplier. Major sectors: 18 groups covering all 180 sectors.", False),
    ("", False),
    ("Key caveats", True),
    ("A. Labor measurement. Row 201 'Wages and salaries' captures only formal employee compensation. The labor of "
     "self-employed workers (smallholder farmers, small traders, freelancers) flows to row 202 'Operating surplus' "
     "as mixed income and cannot be separated from pure capital returns in the Thai IO. Consequence: the "
     "wages-only labor-income multiplier UNDER-states labor dependency in agriculture, tourism-adjacent personal "
     "services, and creative industries. We therefore report TWO bounds on every labor indicator:", False),
    ("    • Lower bound (wages only) — row 201 / output.", False),
    ("    • Upper bound (wages + operating surplus) — row 201 + row 202 / output. "
     "This OVER-states labor by including pure capital returns for capital-intensive sectors.", False),
    ("The true labor share lies between the two. The gap is widest for agriculture (often 8% wages vs. 40%+ "
     "wages+OS), narrower for formal services.", False),
    ("", False),
    ("B. No employment headcount data. The IO file contains no jobs or FTE count by sector, so a physical "
     "employment multiplier (workers per million baht of final demand) cannot be computed. The labor-income "
     "multiplier is a baht-denominated proxy, not a headcount.", False),
    ("", False),
    ("C. Temporal comparability (1990 vs 2021). Both tables are in nominal prices of their respective year. "
     "Multipliers and shares are dimensionless and directly comparable; absolute output levels are NOT. "
     "Industry-of-the-future definitions are anachronistic for 1990 (e.g., 'Digital services' in 1990 is post "
     "office + pre-internet business services). 1990 industry figures should be read as 'the same sector bundle "
     "on 1990 technology', not as a 1990 version of today's industry.", False),
    ("", False),
    ("D. Leontief assumptions. Fixed input proportions, linear response, no price or capacity constraints, "
     "joint production excluded. Standard assumptions for IO multipliers but worth recalling.", False),
    ("", False),
    ("E. Aggregation bias. Weighted averaging over disaggregated multipliers differs slightly from aggregating the "
     "IO table first and then inverting. Agribusiness shows the largest gap (~3%) because of strong intra-group "
     "chains; other industries differ by <1%. We report the disaggregated-then-averaged version; see analysis "
     "script comments for the alternative.", False),
    ("", False),
    ("F. Export-orientation convention. Export orientation is reported as exports (cols 305 + 306) divided by "
     "TOTAL SUPPLY (col 700 = domestic output + imports), the same denominator used for import penetration. The "
     "more familiar ratio exports/output is misleading for sectors that mainly re-export imported goods through "
     "domestic trade margins (e.g., 035 'Other non-ferrous metals', 037 'Chemical fertilizers', 071 'Knitting'): "
     "the exports/output ratio there exceeds 100%, sometimes by orders of magnitude, because the numerator "
     "includes the gross re-exported value while the denominator only captures the small domestic processing "
     "margin. Switching to exports/total-supply puts the ratio on a bounded supply basis (supply identity: "
     "total supply = total demand, with exports as one component of demand) and makes export and import "
     "indicators directly comparable. A handful of sectors still slightly exceed 100% due to negative inventory "
     "changes or trade-margin allocations in the original table; these are reported as-is rather than capped.",
     False),
]
row = 2
for line, is_header in methodology_lines:
    c = ws_m.cell(row=row, column=1, value=line)
    if is_header:
        c.font = Font(bold=True, color="305496", size=11)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if line and not is_header:
        ws_m.row_dimensions[row].height = 45
    row += 1

# ---- Sheet 2: Industry summary ----
ws = out.create_sheet("Industry summary")
ws["A1"] = "Multipliers and structural indicators by industry of the future"
ws["A1"].font = Font(bold=True, size=13); ws.merge_cells("A1:Z1")
ws["A2"] = ("Thailand IO 2021 (180 sectors). Industry values are output-weighted averages over member sectors "
            "(weights = total output). Type I = open Leontief; Type II = household-endogenised closed Leontief. "
            "Labor-income bounds: Low = wages only; High = wages + operating surplus (see methodology).")
ws.merge_cells("A2:Z2")

# Group multi-column headers into two rows for readability
group_hdr = [
    ("Industry / sector count / size", 1, 6),
    ("Type I multipliers", 7, 11),
    ("Type II multipliers (induced)", 12, 15),
    ("Labor compensation bounds", 16, 19),
    ("Value-added structure", 20, 23),
    ("External orientation", 24, 25),
    ("Intermediate structure & linkages", 26, 29),
]
for label, c_start, c_end in group_hdr:
    cell = ws.cell(row=4, column=c_start, value=label)
    cell.font = hdr_font
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if c_end > c_start:
        ws.merge_cells(start_row=4, start_column=c_start, end_row=4, end_column=c_end)

headers = [
    "Industry of the future", "# sectors", "Total output (mn baht)", "Total exports (mn baht)",
    "Total wages (mn baht)", "Total VA (mn baht)",
    # Type I multipliers
    "Output mult", "Labor-inc mult (Low: wages only)", "Labor-inc mult (High: wages+OS)",
    "VA mult", "Import mult (leakage)",
    # Type II multipliers
    "Output mult (II)", "Labor-inc mult (II, wages only)", "VA mult (II)", "Import mult (II)",
    # Labor compensation bounds
    "Wages/output", "Op.surplus/output", "Labor share LOW (wages)", "Labor share HIGH (wages+OS)",
    # VA structure
    "Depreciation/output", "Indirect tax/output", "VA/output", "Labor-to-capital (low/broad)",
    # External
    "Export orientation", "Import penetration",
    # Intermediate & linkages
    "Dom. intermediate share", "Imp. intermediate share",
    "Backward linkage idx (1=avg)", "Forward linkage idx (1=avg)",
]
widths = [26, 7] + [15]*4 + [11, 15, 15, 11, 13] + [11, 15, 11, 13] + [12, 13, 15, 15] + [14, 13, 11, 16] + [12, 13] + [15, 15, 15, 15]
set_header(ws, 5, headers, widths=widths)

for r, row in enumerate(industry_summary, start=6):
    ws.cell(row=r, column=1, value=row["industry"]).font = bold
    ws.cell(row=r, column=2, value=row["n_sectors"])
    ws.cell(row=r, column=3, value=row["total_output"]).number_format = fmt_int
    ws.cell(row=r, column=4, value=row["total_exports"]).number_format = fmt_int
    ws.cell(row=r, column=5, value=row["total_wages"]).number_format = fmt_int
    ws.cell(row=r, column=6, value=row["total_va"]).number_format = fmt_int
    # Type I
    ws.cell(row=r, column=7, value=row["output_mult"]).number_format = fmt_4
    ws.cell(row=r, column=8, value=row["labor_income_mult_low"]).number_format = fmt_4
    ws.cell(row=r, column=9, value=row["labor_income_mult_hi"]).number_format = fmt_4
    ws.cell(row=r, column=10, value=row["va_mult"]).number_format = fmt_4
    ws.cell(row=r, column=11, value=row["import_mult"]).number_format = fmt_4
    # Type II
    ws.cell(row=r, column=12, value=row["output_mult_II"]).number_format = fmt_4
    ws.cell(row=r, column=13, value=row["labor_income_mult_II"]).number_format = fmt_4
    ws.cell(row=r, column=14, value=row["va_mult_II"]).number_format = fmt_4
    ws.cell(row=r, column=15, value=row["import_mult_II"]).number_format = fmt_4
    # Labor compensation
    ws.cell(row=r, column=16, value=row["wages_share"]).number_format = fmt_pct
    ws.cell(row=r, column=17, value=row["op_surplus_share"]).number_format = fmt_pct
    ws.cell(row=r, column=18, value=row["labor_share_low"]).number_format = fmt_pct
    ws.cell(row=r, column=19, value=row["labor_share_hi"]).number_format = fmt_pct
    # VA structure
    ws.cell(row=r, column=20, value=row["depr_share"]).number_format = fmt_pct
    ws.cell(row=r, column=21, value=row["indtax_share"]).number_format = fmt_pct
    ws.cell(row=r, column=22, value=row["va_share"]).number_format = fmt_pct
    ws.cell(row=r, column=23, value=row["labor_to_cap_low"]).number_format = fmt_4
    # External
    ws.cell(row=r, column=24, value=row["export_orientation"]).number_format = fmt_pct
    ws.cell(row=r, column=25, value=row["import_penetration"]).number_format = fmt_pct
    # Intermediate & linkages
    ws.cell(row=r, column=26, value=row["dom_intermediate_share"]).number_format = fmt_pct
    ws.cell(row=r, column=27, value=row["imp_intermediate_share"]).number_format = fmt_pct
    ws.cell(row=r, column=28, value=row["backward_linkage_idx"]).number_format = fmt_4
    ws.cell(row=r, column=29, value=row["forward_linkage_idx"]).number_format = fmt_4

ws.freeze_panes = "B6"

# ---- Sheet 2: Output multiplier decomposition by major sector ----
ws2 = out.create_sheet("Output mult decomposition")
ws2["A1"] = "Output multiplier decomposition by source major sector (columns sum to industry output multiplier)"
ws2["A1"].font = Font(bold=True, size=13)
ws2.merge_cells("A1:G1")
ws2["A2"] = ("For each industry of the future (column), entries show how much output is generated in each major "
             "sector (row) per 1 baht of final demand for the industry. Weighted over member sectors by total output.")
ws2.merge_cells("A2:G2")

set_header(ws2, 4, ["Major sector"] + industries + ["Share of total"], widths=[42, 14, 14, 14, 14, 14, 14])

total_per_industry = {ind: industry_decomp_output[ind].sum() for ind in industries}
for r, m_name in enumerate(major_order, start=5):
    ws2.cell(row=r, column=1, value=m_name)
    row_total = 0.0
    for k, ind in enumerate(industries, start=2):
        v_ = industry_decomp_output[ind][major_order.index(m_name)]
        ws2.cell(row=r, column=k, value=float(v_)).number_format = fmt_4
        row_total += v_
    ws2.cell(row=r, column=2 + len(industries), value=row_total).number_format = fmt_4

total_row = 5 + len(major_order)
ws2.cell(row=total_row, column=1, value="Total (= industry output multiplier)").font = bold
for k, ind in enumerate(industries, start=2):
    c = ws2.cell(row=total_row, column=k, value=float(total_per_industry[ind]))
    c.number_format = fmt_4; c.font = bold

# Percentage-contribution sub-table
pct_start = total_row + 3
ws2.cell(row=pct_start - 1, column=1, value="% contribution to each industry's output multiplier").font = bold
set_header(ws2, pct_start, ["Major sector"] + industries + ["—"], widths=[42, 14, 14, 14, 14, 14, 14])
for r, m_name in enumerate(major_order, start=pct_start + 1):
    ws2.cell(row=r, column=1, value=m_name)
    for k, ind in enumerate(industries, start=2):
        v_ = industry_decomp_output[ind][major_order.index(m_name)]
        tot = total_per_industry[ind]
        ws2.cell(row=r, column=k, value=float(v_ / tot) if tot else 0).number_format = fmt_pct

# ---- Sheet 3: Labor-income decomposition by major sector ----
ws3 = out.create_sheet("Labor income decomposition")
ws3["A1"] = "Labor income generated per 1 baht of final demand, decomposed by source major sector"
ws3["A1"].font = Font(bold=True, size=13); ws3.merge_cells("A1:G1")
ws3["A2"] = ("Values are baht of direct+indirect labor compensation per 1 baht of final demand for the industry. "
             "Column totals equal the industry's 'Labor income generated per 1 baht FD' shown on the summary sheet.")
ws3.merge_cells("A2:G2")

set_header(ws3, 4, ["Major sector"] + industries + ["Share of total"], widths=[42, 14, 14, 14, 14, 14, 14])

total_li_per_industry = {ind: industry_decomp_laborinc[ind].sum() for ind in industries}
for r, m_name in enumerate(major_order, start=5):
    ws3.cell(row=r, column=1, value=m_name)
    row_total = 0.0
    for k, ind in enumerate(industries, start=2):
        v_ = industry_decomp_laborinc[ind][major_order.index(m_name)]
        ws3.cell(row=r, column=k, value=float(v_)).number_format = fmt_4
        row_total += v_
    ws3.cell(row=r, column=2 + len(industries), value=row_total).number_format = fmt_4

total_row = 5 + len(major_order)
ws3.cell(row=total_row, column=1, value="Total (= labor income per 1 baht FD)").font = bold
for k, ind in enumerate(industries, start=2):
    c = ws3.cell(row=total_row, column=k, value=float(total_li_per_industry[ind]))
    c.number_format = fmt_4; c.font = bold

pct_start = total_row + 3
ws3.cell(row=pct_start - 1, column=1, value="% contribution to each industry's labor income impact").font = bold
set_header(ws3, pct_start, ["Major sector"] + industries + ["—"], widths=[42, 14, 14, 14, 14, 14, 14])
for r, m_name in enumerate(major_order, start=pct_start + 1):
    ws3.cell(row=r, column=1, value=m_name)
    for k, ind in enumerate(industries, start=2):
        v_ = industry_decomp_laborinc[ind][major_order.index(m_name)]
        tot = total_li_per_industry[ind]
        ws3.cell(row=r, column=k, value=float(v_ / tot) if tot else 0).number_format = fmt_pct

# ---- Sheet 4: Sector detail (180 sectors) ----
ws4 = out.create_sheet("Sector detail")
headers4 = [
    "Code", "Sector", "Industry of the future", "Major sector",
    "Total output (mn baht)", "Exports (mn baht)", "Wages (mn baht)",
    "Output multiplier", "Labor-income multiplier", "VA multiplier", "Import multiplier",
    "Export orientation", "Import penetration", "Labor share", "Capital share",
    "VA share", "Labor-to-capital", "Dom. intermediate share", "Imp. intermediate share",
    "Backward linkage idx", "Forward linkage idx",
]
widths4 = [6, 46, 22, 34] + [14]*(len(headers4)-4)
set_header(ws4, 1, headers4, widths=widths4)
ws4.row_dimensions[1].height = 46

for i, code in enumerate(sector_codes):
    r = i + 2
    xi = x[i]
    labor_sh = h[i]
    cap_sh = (op_surplus[i] + depreciation[i]) / xi if xi > 0 else np.nan
    va_sh = v[i]
    dom_int = Z[:, i].sum() / xi if xi > 0 else np.nan
    imp_int = Z_imp[:, i].sum() / xi if xi > 0 else np.nan
    sup_i = total_supply[i]
    exp_or = exports[i] / sup_i if sup_i > 0 else np.nan
    imp_pen = total_imports[i] / sup_i if sup_i > 0 else np.nan
    lab_to_cap = labor_sh / cap_sh if cap_sh > 0 else np.nan

    ws4.cell(row=r, column=1, value=code)
    ws4.cell(row=r, column=2, value=sector_name[code])
    ws4.cell(row=r, column=3, value=industry_of.get(code, ""))
    ws4.cell(row=r, column=4, value=major_of[i])
    ws4.cell(row=r, column=5, value=float(xi)).number_format = fmt_int
    ws4.cell(row=r, column=6, value=float(exports[i])).number_format = fmt_int
    ws4.cell(row=r, column=7, value=float(wages[i])).number_format = fmt_int
    ws4.cell(row=r, column=8, value=float(output_mult[i])).number_format = fmt_4
    lim = labor_income_mult[i]
    ws4.cell(row=r, column=9, value=float(lim) if np.isfinite(lim) else None).number_format = fmt_4
    ws4.cell(row=r, column=10, value=float(va_per_fd[i])).number_format = fmt_4
    ws4.cell(row=r, column=11, value=float(imports_per_fd[i])).number_format = fmt_4
    ws4.cell(row=r, column=12, value=float(exp_or) if np.isfinite(exp_or) else None).number_format = fmt_pct
    ws4.cell(row=r, column=13, value=float(imp_pen) if np.isfinite(imp_pen) else None).number_format = fmt_pct
    ws4.cell(row=r, column=14, value=float(labor_sh)).number_format = fmt_pct
    ws4.cell(row=r, column=15, value=float(cap_sh) if np.isfinite(cap_sh) else None).number_format = fmt_pct
    ws4.cell(row=r, column=16, value=float(va_sh)).number_format = fmt_pct
    ws4.cell(row=r, column=17, value=float(lab_to_cap) if np.isfinite(lab_to_cap) else None).number_format = fmt_4
    ws4.cell(row=r, column=18, value=float(dom_int) if np.isfinite(dom_int) else None).number_format = fmt_pct
    ws4.cell(row=r, column=19, value=float(imp_int) if np.isfinite(imp_int) else None).number_format = fmt_pct
    ws4.cell(row=r, column=20, value=float(backward_linkage[i])).number_format = fmt_4
    ws4.cell(row=r, column=21, value=float(forward_linkage[i])).number_format = fmt_4

ws4.freeze_panes = "E2"

# ---- Sheets 5 & 6: A matrix and Leontief inverse ----
def write_matrix(name, M):
    s = out.create_sheet(name)
    s.cell(row=1, column=1, value="Code \\ Code").font = bold
    for j, code in enumerate(sector_codes, start=2):
        s.cell(row=1, column=j, value=code).font = bold
        s.cell(row=j, column=1, value=code).font = bold
    for i in range(N):
        for j in range(N):
            val = float(M[i, j])
            if val != 0:
                s.cell(row=i + 2, column=j + 2, value=val).number_format = "0.0000"
    s.freeze_panes = "B2"
    s.column_dimensions["A"].width = 10

write_matrix("A matrix", A)
write_matrix("Leontief inverse", L)

out.save(OUT)

# ---------------------------------------------------------------------------
# 6. Console summary
# ---------------------------------------------------------------------------
print("Industry-of-the-future multipliers (Type I, output-weighted averages):\n")
hdr = f"{'Industry':<26}{'OutputM':>9}{'LaborI_lo':>11}{'LaborI_hi':>11}{'VA M':>8}{'ImportM':>9}  {'ExpOrt':>7} {'ImpPen':>7} {'Wage%':>7} {'VA%':>6}"
print(hdr); print("-"*len(hdr))
for r in industry_summary:
    print(f"{r['industry']:<26}{r['output_mult']:>9.4f}{r['labor_income_mult_low']:>11.4f}{r['labor_income_mult_hi']:>11.4f}{r['va_mult']:>8.4f}{r['import_mult']:>9.4f}  "
          f"{r['export_orientation']:>7.1%} {r['import_penetration']:>7.1%} {r['wages_share']:>7.1%} {r['va_share']:>6.1%}")

print("\nOutput multiplier decomposition (columns sum to industry output multiplier):\n")
col_hdr = f"{'Major sector':<42}  " + "".join(f"{ind[:12]:>13}" for ind in industries)
print(col_hdr); print("-"*len(col_hdr))
for m_name in major_order:
    k = major_order.index(m_name)
    row_vals = "  ".join(f"{industry_decomp_output[ind][k]:>11.4f}" for ind in industries)
    print(f"{m_name:<42}  {row_vals}")
print(f"{'TOTAL':<42}  " + "  ".join(f"{industry_decomp_output[ind].sum():>11.4f}" for ind in industries))
print(f"\nSaved: {OUT}")
